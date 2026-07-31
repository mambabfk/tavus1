#!/usr/bin/env python3
"""Write validated answers into a copy of the original SAP form.

Refuses to run if validate.py fails. Never edits the blank original.
Produces a gap report alongside the filled workbook.

Usage: python scripts/fill_workbook.py --out output/SAP_DPP_filled.xlsx
"""
import argparse, json, glob, os, shutil, subprocess, sys
import openpyxl
from openpyxl.styles import Font

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SPEC = json.load(open(os.path.join(ROOT, "reference/form_spec.json")))
ROW_OF = {r["id"]: r["row"] for r in SPEC["rows"] if r["id"]}
ALLOWED = {r["id"]: r["allowed_values"] for r in SPEC["rows"] if r["allowed_values"]}
BLANK = os.path.join(ROOT, "reference/blank_form.xlsx")


def load_program_answers():
    import yaml
    return yaml.safe_load(open(os.path.join(ROOT, "config/tavus_program_answers.yaml"))) or {}


def load_overflow_handling():
    import yaml
    cfg = yaml.safe_load(open(os.path.join(ROOT, "config/scope.yaml"))) or {}
    return cfg.get("overflow_handling")


def program_map(cfg):
    """config keys are '<row_id>_<slug>' -> map back to row ids."""
    out = {}
    for k, v in cfg.items():
        rid = k.split("_", 1)[0]
        if rid in ROW_OF:
            out[rid] = v
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="output/SAP_DPP_filled.xlsx")
    ap.add_argument("--skip-validate", action="store_true",
                    help="Only for debugging. Never use on a file you will submit.")
    args = ap.parse_args()

    if not args.skip_validate:
        r = subprocess.run([sys.executable, os.path.join(ROOT, "scripts/validate.py")])
        if r.returncode != 0:
            print("\nvalidate.py failed. Nothing written.")
            sys.exit(1)

    out = os.path.join(ROOT, args.out)
    os.makedirs(os.path.dirname(out), exist_ok=True)

    files = sorted(glob.glob(os.path.join(ROOT, "data/verified/*.json"))) or \
            sorted(glob.glob(os.path.join(ROOT, "data/vendors/*.json")))
    cols = list(SPEC["vendor_columns"].values())
    if len(files) > len(cols) and load_overflow_handling() != "second_sheet":
        print(f"{len(files)} vendors but {len(cols)} columns and scope.yaml's "
              "overflow_handling is not 'second_sheet'. Resolve scope first.")
        sys.exit(1)

    prog = program_map(load_program_answers())
    gaps, sourced = [], 0
    outputs = []

    base, ext = os.path.splitext(out)
    for sheet_no, start in enumerate(range(0, len(files), len(cols)), 1):
        chunk = files[start:start + len(cols)]
        out_path = out if sheet_no == 1 else f"{base}_{sheet_no}{ext}"
        shutil.copy(BLANK, out_path)
        wb = openpyxl.load_workbook(out_path)
        ws = wb[SPEC["sheet"]]

        for col, path in zip(cols, chunk):
            d = json.load(open(path))
            name = d.get("vendor_name") or os.path.basename(path)
            ws[f"{col}1"] = name

            for rid, val in prog.items():          # Bucket B, uniform
                if rid in ALLOWED and val not in ALLOWED[rid]:
                    continue  # conditional dropdown (e.g. 1.5.3 when 1.5=No): leave empty
                ws[f"{col}{ROW_OF[rid]}"] = val

            for rid, f in (d.get("fields") or {}).items():
                if rid not in ROW_OF:
                    gaps.append(f"{name}: unknown field id '{rid}', skipped")
                    continue
                val = f.get("value")
                if val in (None, ""):
                    continue
                if rid in ALLOWED and val not in ALLOWED[rid]:
                    gaps.append(f"{name}/{rid}: illegal dropdown value '{val}', left blank")
                    continue
                cell = ws[f"{col}{ROW_OF[rid]}"]
                cell.value = val
                if val == "Not known":
                    cell.font = Font(name="Arial", italic=True)
                    gaps.append(f"{name}/{rid}: NOT KNOWN — {f.get('notes','no note')}")
                else:
                    sourced += 1

        wb.save(out_path)
        outputs.append(out_path)

    report = os.path.join(os.path.dirname(out), "gap_report.md")
    with open(report, "w") as fh:
        fh.write("# Gap report — must be closed before submitting to SAP\n\n")
        fh.write(f"{sourced} cells filled with sourced values. "
                 f"{len(gaps)} items need attention.\n\n")
        for g in gaps:
            fh.write(f"- {g}\n")
        fh.write("\n## Reminders\n"
                 "- Legal entity names must match the TOM questionnaire #8.3 submission.\n"
                 "- Row 1.6 (incidents) and Section 2 (data categories) are counsel's call.\n"
                 "- Row 1.7: answering Yes invites SAP to require non-US processing.\n")

    for p in outputs:
        print(f"Wrote {p}")
    print(f"Wrote {report}  ({sourced} sourced cells, {len(gaps)} gaps)")
    print("\nThis form is not done. Read the gap report.")


if __name__ == "__main__":
    main()
